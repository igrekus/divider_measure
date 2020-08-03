import datetime
import openpyxl
import os

from PyQt5 import uic
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QWidget


from measureresult import phs_value_for_phs_code, att_value_for_att_code, chunks


class StatWidget(QWidget):

    def __init__(self, parent=None, result=None):
        super().__init__(parent)

        self._out_dir = 'xlsx'
        self._result = result

        self._ui = uic.loadUi('statwidget.ui', self)

        self._ui.texteditStat.setPlainText('')

    @pyqtSlot()
    def on_btnExportExcel_clicked(self):
        if not os.path.isdir(self._out_dir):
            os.mkdir(self._out_dir)

        self._exportExcel()

    @property
    def stats(self):
        return self._ui.texteditStat.plainText()

    @stats.setter
    def stats(self, text):
        self._ui.texteditStat.setPlainText(text)

    def _exportExcel(self):

        def build_title(titles):
            cells = []
            for title in titles:
                for att, phs in title:
                    cells.append(f'{att} dB, {phs}°')

                cells.append('')
                cells.append('F, GHz')
            return ['F, GHz'] + cells[:-1]

        def build_data(freqs, ampss):
            rows = [list() for _ in range(len(freqs))]
            for amps in ampss:
                for row, freq, *amp in zip(rows, freqs, *amps):
                    row += [freq] + amp + ['']
            return rows

        unique_phases = set(self._result._phase_codes)
        chunk_len = len(unique_phases)

        title_data = [(att_value_for_att_code(att), phs_value_for_phs_code(phs)) for att, phs in zip(self._result._att_codes, self._result._phase_codes)]
        title_data = [chunk for chunk in chunks(title_data, chunk_len)]

        freqs = self._result.freqs
        s21_amps = [chunk for chunk in chunks(self._result.s21, chunk_len)]
        s21_phs = [chunk for chunk in chunks(self._result.phase, chunk_len)]
        s11_amps = [chunk for chunk in chunks(self._result.s11, chunk_len)]
        s12_amps = [chunk for chunk in chunks(self._result.s12, chunk_len)]
        s22_amps = [chunk for chunk in chunks(self._result.s22, chunk_len)]
        titles = build_title(title_data)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = 'S21 amps'
        ws.append(titles)
        for row in build_data(freqs, s21_amps):
            ws.append(row)

        ws = wb.create_sheet('S21 phases')
        ws.append(titles)
        for row in build_data(freqs, s21_phs):
            ws.append(row)

        ws = wb.create_sheet('S11 amps')
        ws.append(titles)
        for row in build_data(freqs, s11_amps):
            ws.append(row)

        ws = wb.create_sheet('S12 amps')
        ws.append(titles)
        for row in build_data(freqs, s12_amps):
            ws.append(row)

        ws = wb.create_sheet('S22 amps')
        ws.append(titles)
        for row in build_data(freqs, s22_amps):
            ws.append(row)

        if not os.path.isdir('xlsx'):
            os.mkdir('xlsx')

        wb.save(f'xlsx\\{self.xlsx_filename}')
        os.startfile('xlsx')

    @property
    def xlsx_filename(self):
        return f'att-psm_{datetime.datetime.now().isoformat().replace(":","-")}.xlsx'
